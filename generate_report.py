#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BUSINESSES = [
    {
        "שם העסק": 'מכללת ש"ש-האקדמיה לשיווק שותפים',
        "כתובת": "איסלנד 17, ירושלים",
        "סוג עסק": "קורסים והשתלמויות",
        "קישור 1": "https://www.d.co.il/80181095/8190/",
    },
    {
        "שם העסק": "רונית שקלים",
        "כתובת": "איסלנד 17, ירושלים",
        "סוג עסק": "גרפיקאים וסטודיו לגרפיקה",
        "קישור 1": "https://www.d.co.il/80027963/11301/",
    },
    {
        "שם העסק": "מוטי רבינוביץ מאמן כושר אישי",
        "כתובת": "איסלנד 21, ירושלים",
        "סוג עסק": "מדריכי ספורט וכושר",
        "קישור 1": "https://www.d.co.il/80272956/22570/",
    },
    {
        "שם העסק": "יעל פרץ-מורה לצרפתית",
        "כתובת": "אלעזר המודעי 12, ירושלים",
        "סוג עסק": "שיעורים פרטיים",
        "קישור 1": "https://www.d.co.il/80015963/24630/",
    },
    {
        "שם העסק": 'גלעד גלי - שרה עו"ד',
        "כתובת": "אלפסי 21, ירושלים",
        "סוג עסק": "עורכי דין פלילי",
        "קישור 1": "https://www.d.co.il/27624410/9010106/",
    },
    {
        "שם העסק": 'עו"ד חוה ולר',
        "כתובת": "אלפנדרי 36, ירושלים",
        "סוג עסק": "עורכי דין - דיני מקרקעין",
        "קישור 1": "https://www.d.co.il/80231621/9010118/",
    },
    {
        "שם העסק": "מזיקיס-המדביר היווני",
        "כתובת": "אלפנדרי 6, ירושלים",
        "סוג עסק": "הדברה",
        "קישור 1": "https://www.d.co.il/80182149/13110/",
    },
    {
        "שם העסק": 'המכון לעסקאות עתידיות ואופציות בע"מ',
        "כתובת": "אלפסי 23, ירושלים",
        "סוג עסק": "יועצי השקעות",
        "קישור 1": "https://www.d.co.il/25760070/14400/",
    },
    {
        "שם העסק": "שטרנגר גבי",
        "כתובת": "אלקחי מרדכי 20, ירושלים",
        "סוג עסק": "פסיכולוגים קליניים",
        "קישור 1": "https://www.d.co.il/60148150/40270/",
    },
    {
        "שם העסק": "דריבן שירותי אבטחה ואימונים",
        "כתובת": "אלקחי מרדכי 38, ירושלים",
        "סוג עסק": "שמירה ואבטחה",
        "קישור 1": "https://www.d.co.il/80014328/49230/",
    },
    {
        "שם העסק": "שרית מהגר אמנות האיפור",
        "כתובת": "אלקחי מרדכי 8, ירושלים",
        "סוג עסק": "איפור",
        "קישור 1": "https://www.d.co.il/80054059/2800/",
    },
    {
        "שם העסק": 'עו"ד גבאי דניאל',
        "כתובת": "אמיל זולא 6, ירושלים",
        "סוג עסק": "עורכי דין - דיני מקרקעין",
        "קישור 1": "https://www.d.co.il/80276756/9010118/",
    },
    {
        "שם העסק": "ארנסקי אבי",
        "כתובת": "אמיל זולא 6, ירושלים",
        "סוג עסק": "תרגום",
        "קישור 1": "https://www.d.co.il/75711850/51870/",
    },
    {
        "שם העסק": "יתד התחדשות עירונית",
        "כתובת": "אמיל זולא 6, ירושלים",
        "סוג עסק": "נדל\"ן - התחדשות עירונית",
        "קישור 1": "https://www.d.co.il/80243217/34590/",
    },
    {
        "שם העסק": "ענת לביא",
        "כתובת": "אמציה 8, ירושלים",
        "סוג עסק": "פסיכותרפיה",
        "קישור 1": "https://www.d.co.il/80102167/40230/",
    },
    {
        "שם העסק": "לוי יוסי הובלות והרכבות",
        "כתובת": "אנטיגנוס 11, ירושלים",
        "סוג עסק": "חברות הובלה ומובילים",
        "קישור 1": "https://www.d.co.il/80032602/35840/",
    },
    {
        "שם העסק": "גולן עליזה",
        "כתובת": "אנילביץ 64, ירושלים",
        "סוג עסק": "פלדנקרייז - טיפול בתנועה",
        "קישור 1": "https://www.d.co.il/80065084/40030014/",
    },
    {
        "שם העסק": "גן האור - גן אנתרופוסופי",
        "כתובת": "אנילביץ 26, ירושלים",
        "סוג עסק": "גני ילדים",
        "קישור 1": "https://www.d.co.il/80108035/10880/",
    },
    {
        "שם העסק": "כהן אמנון",
        "כתובת": "אנילביץ 68, ירושלים",
        "סוג עסק": "מורה לנהיגה",
        "קישור 1": "https://www.d.co.il/24466980/8450/",
    },
    {
        "שם העסק": "נלפ ישראל",
        "כתובת": "אסתר המלכה 2, ירושלים",
        "סוג עסק": "אבחון וטיפול בלקויי למידה",
        "קישור 1": "https://www.d.co.il/80166202/22310/",
    },
]

HEADERS = ["שם העסק", "כתובת", "סוג עסק", "דירוג אינדיקציה", "קישור 1", "קישור 2", "קישור 3"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "דוח נכסים"
ws.sheet_view.rightToLeft = True

H_BG = "1F4E79"
H_F = "FFFFFF"

def thin_border():
    s = Side(style="thin", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

col_widths = [36, 34, 28, 18, 50, 50, 50]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Header row
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color=H_F, name="Arial", size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor=H_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

ws.row_dimensions[1].height = 22

ROW_COLORS = ["FFFFFF", "D6E4F0"]

for row_idx, biz in enumerate(BUSINESSES, 2):
    row_color = ROW_COLORS[(row_idx - 2) % 2]
    fill = PatternFill(fill_type="solid", fgColor=row_color)

    values = [
        biz.get("שם העסק", ""),
        biz.get("כתובת", ""),
        biz.get("סוג עסק", ""),
        "גבוה",
        biz.get("קישור 1", ""),
        biz.get("קישור 2", ""),
        biz.get("קישור 3", ""),
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.border = thin_border()
        cell.font = Font(name="Arial", size=10)

        if col_idx >= 5 and val and val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    ws.row_dimensions[row_idx].height = 18

filename = "דוח נכסים לתאריך:21-07-2026.xlsx"
filepath = f"/home/user/Final_Project/{filename}"
wb.save(filepath)
print(f"Saved: {filepath}")
print(f"Total businesses: {len(BUSINESSES)}")
